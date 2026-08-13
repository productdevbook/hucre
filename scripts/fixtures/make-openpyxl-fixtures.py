#!/usr/bin/env python3
"""Regenerate the openpyxl half of test/fixtures/.

    python3 scripts/fixtures/make-openpyxl-fixtures.py test/fixtures

#464 asks for files hucre did not write. The Excel corpus answered that
with one producer; this adds a second, and a second producer is not a
formality. Excel is a single point of agreement — if hucre and Excel
happen to share a misunderstanding, a corpus made only of Excel output
cannot see it. openpyxl was written by different people from a different
reading of the same spec, and it emits four shapes Excel never does:

  * ``<f>A1*3</f><v/>`` — a formula with no cached result, and an *empty*
    value element rather than an absent one. Excel always caches.
  * ``t="d"`` cells holding an ISO 8601 date instead of a serial number
    (ECMA-376 ST_CellType 'd'). Excel never writes these.
  * ``date1904="1"``. Reachable in Excel only through a workbook setting
    most Windows users never touch.
  * ``t="inlineStr"`` with ``<is><t>`` and no shared string table at all,
    in write-only mode. Excel always writes ``sharedStrings.xml``.

``openpyxl-basic.xlsx`` deliberately mirrors ``excel-basic.xlsx``, so the
two producers can be read into the same model and compared;
``test/real-files.test.ts`` asserts they agree.

Determinism: document timestamps are pinned, so regenerating differs only
in ZIP entry times, and creator/lastModifiedBy are blanked to match the
guarantee the Excel fixtures make (see test/fixtures/PROVENANCE.md).

Developed against openpyxl 3.1.5. CI does not run this.
"""

from __future__ import annotations

import datetime
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.datetime import CALENDAR_MAC_1904

# Pinned so regenerating does not churn docProps/core.xml.
STAMP = datetime.datetime(2024, 1, 1, 0, 0, 0)

DATE = datetime.date(2024, 3, 17)


def new_book(title: str, write_only: bool = False) -> openpyxl.Workbook:
    wb = openpyxl.Workbook(write_only=write_only)
    if write_only:
        wb.create_sheet(title)
    else:
        wb.active.title = title
    # The Excel corpus guarantees no author survives into a fixture;
    # openpyxl stamps "openpyxl" as the creator, so blank it for the same
    # reason and so one assertion covers every file in the directory.
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.created = STAMP
    wb.properties.modified = STAMP
    return wb


def basic(path: Path) -> None:
    """The value cells of excel-basic.xlsx, from the other producer.

    Column E holds the numbers Excel cached for its formulas, written as
    literal values: openpyxl does not evaluate formulas, so putting
    ``=B2*2`` here would compare a cached result against nothing. The
    formula-without-a-result case is its own fixture below, where it is
    the point rather than a confound.
    """
    wb = new_book("Data")
    ws = wb.active
    for col, head in enumerate(["Name", "Qty", "Date", "Active", "Total"], start=1):
        ws.cell(row=1, column=col, value=head)

    rows = [
        ("Widget", 12, datetime.date(2024, 3, 17), True, 24),
        ("Gadget", -3.5, datetime.date(1999, 12, 31), False, -7),
        ("Doohickey", 0, datetime.date(2024, 2, 29), True, 8.5),
    ]
    for r, (name, qty, date, active, total) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=qty)
        cell = ws.cell(row=r, column=3, value=date)
        cell.number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4, value=active)
        ws.cell(row=r, column=5, value=total)
    wb.save(path)


def formulas(path: Path) -> None:
    """Formulas with no cached result — the shape Excel cannot produce."""
    wb = new_book("Formulas")
    ws = wb.active
    ws["A1"] = 2
    ws["A2"] = 3
    ws["B1"] = "=A1*3"
    ws["B2"] = "=SUM(A1:A2)"
    ws["B3"] = '=CONCATENATE("x","y")'
    # A formula that would be an error if anything ever evaluated it.
    ws["B4"] = "=1/0"
    # A plain value after the formulas, so a reader that loses its place
    # in the row does not simply return a short row and look fine.
    ws["C1"] = "after"
    wb.save(path)


def isodates(path: Path) -> None:
    """``t="d"`` cells: the date as ISO 8601 text, not a serial number.

    Row 4 is a date written the ordinary way — a serial number under a
    date format — so the fixture carries its own control.
    """
    wb = new_book("IsoDates")
    wb.iso_dates = True
    ws = wb.active
    ws["A1"] = "date"
    ws["B1"] = DATE
    ws["A2"] = "datetime"
    ws["B2"] = datetime.datetime(2024, 3, 17, 13, 45, 30)
    ws["A3"] = "time"
    ws["B3"] = datetime.time(13, 45, 30)
    ws["A4"] = "serial control"
    control = ws.cell(row=4, column=2, value=45368)
    control.number_format = "yyyy-mm-dd"
    wb.save(path)


def epoch1904(path: Path) -> None:
    """``date1904="1"``.

    Every serial here means a different day if the flag is ignored — the
    two epochs are 1,462 days apart — so a reader that drops it cannot
    fail quietly.
    """
    wb = new_book("Epoch1904")
    wb.epoch = CALENDAR_MAC_1904
    ws = wb.active
    values = [
        ("epoch", datetime.date(1904, 1, 1)),
        ("date", DATE),
        ("datetime", datetime.datetime(2024, 3, 17, 13, 45, 30)),
        ("leap day", datetime.date(2024, 2, 29)),
    ]
    for r, (label, value) in enumerate(values, start=1):
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=value)
        cell.number_format = "yyyy-mm-dd hh:mm:ss"
    wb.save(path)


def inline_strings(path: Path) -> None:
    """Write-only mode: ``t="inlineStr"``, and no sharedStrings.xml.

    #441 wanted the whitespace-preservation case from a producer that
    emits inline strings. Excel does not; this does.
    """
    wb = new_book("Inline", write_only=True)
    ws = wb["Inline"]
    ws.append(["  leading", "trailing  "])
    ws.append([" both ", "plain"])
    ws.append(["a & b < c > d", "naïve 日本語 Ωμέγα \U0001f600"])
    ws.append(["line one\nline two", "tab\tsep"])
    ws.append(["  ", 42])
    wb.save(path)


def styled(path: Path) -> None:
    """openpyxl's style serialization, which is not Excel's.

    A different default font, ``t="n"`` written out explicitly on
    numbers, and a styles.xml assembled in a different order — the same
    meaning in different bytes, which is the whole reason a second
    producer is worth having.
    """
    wb = new_book("Styled")
    ws = wb.active

    ws["A1"] = "bold"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "italic 14"
    ws["B1"].font = Font(italic=True, size=14)
    ws["C1"] = "red courier"
    ws["C1"].font = Font(name="Courier New", color="FFFF0000")

    ws["A2"] = "filled"
    ws["A2"].fill = PatternFill(fill_type="solid", start_color="FFFFFF00")
    ws["B2"] = "bordered"
    thin = Side(style="thin")
    ws["B2"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["C2"] = 0.125
    ws["C2"].number_format = "0.00%"

    ws["A3"] = "centred"
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B3"] = "wrapped"
    ws["B3"].alignment = Alignment(wrap_text=True)

    ws["A4"] = "merged"
    ws.merge_cells("A4:C4")
    ws.freeze_panes = "B3"
    wb.save(path)


FIXTURES = {
    "openpyxl-basic.xlsx": basic,
    "openpyxl-formulas.xlsx": formulas,
    "openpyxl-isodates.xlsx": isodates,
    "openpyxl-1904.xlsx": epoch1904,
    "openpyxl-inline-strings.xlsx": inline_strings,
    "openpyxl-styled.xlsx": styled,
}


def main() -> int:
    if len(sys.argv) < 2:
        print(f"usage: {sys.argv[0]} <output-directory>", file=sys.stderr)
        return 2
    out = Path(sys.argv[1])
    out.mkdir(parents=True, exist_ok=True)
    for name, build in FIXTURES.items():
        build(out / name)
        print(f"wrote {out / name}")
    print(f"OK - wrote {len(FIXTURES)} fixtures with openpyxl {openpyxl.__version__}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
